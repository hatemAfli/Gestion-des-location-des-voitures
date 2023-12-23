import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QWidget,QApplication,QDialog,QMainWindow
from PyQt5.uic import *
import datetime
from datetime import timedelta
import webbrowser
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

voiture = {}
client = {}
location= {}
def verife_matr(matr,d):
        for i in d :
            if i==matr :
                return 1
                break
        return 0

class main(QMainWindow):
    
    def __init__(self):
        super(main,self).__init__()
        loadUi("menu_general.ui",self)
        self.actionAjouter_voiture.triggered.connect(self.gotoaddv)
        self.actionSuppression_d_une_voiture_donn_e.triggered.connect(self.suppr_par_matr)
        self.actionSuppression_des_voitures_d_une_marque_donn_e.triggered.connect(self.suppr_par_marque)
        self.actionSuppression_des_voitures_age_10_ans.triggered.connect(self.suppr_sup_10ans)
        self.actionModifier_Prix.triggered.connect(self.modifie_prix)
        self.actionModifier_Couleur.triggered.connect(self.modifie_couleur)
        self.actionContenue_du_dicyionnaire_voitures.triggered.connect(self.affiche_dict_voiture)
        self.actionRecherche_Par_Matricule.triggered.connect(self.affiche_par_matr)
        self.actionRecherche_Par_Couleur.triggered.connect(self.chercher_voit_couleur)
        self.actionRecherche_Par_Marque.triggered.connect(self.chercher_voiture_marque)
        self.actionRecherche_Des_Voitures_Disponibles.triggered.connect(self.chercher_voiture_disp)
        self.actionRecherche_Des_Voitures_Lou_e.triggered.connect(self.chercher_voiture_louee)
        self.actionRecherche_Des_Voitures_Lou_e_Entre_2_Dates.triggered.connect(self.cherche_deux_date)
        self.actionAjouter_Un_Nouvel_Client.triggered.connect(self.ajout_client)
        self.actionSupprimer_Un_Client.triggered.connect(self.supp_client_cin)
        self.actionModifier_Adresse.triggered.connect(self.modifiadresse)
        self.actionModifier_T_l_phone.triggered.connect(self.modifitelephone)
        self.actionModifier_Mail.triggered.connect(self.modifimail)
        self.actionContenus_Du_Dictionnaire_Client.triggered.connect(self.affiche_dict_client)
        self.actionRecherche_Par_CIN.triggered.connect(self.affiche_cin)
        self.actionAjouter_Un_Nouvel_Location.triggered.connect(self.ajout_location)
        self.actionSupprimer_Une_Location.triggered.connect(self.supp_location)
        self.actionModifier_Date_De_Location.triggered.connect(self.mod_date_locat)
        self.actionModifier_Dur_e.triggered.connect(self.mod_duree_locat)
        self.actionContenu_Du_Dictionnaire_Locations.triggered.connect(self.afficher_loc)
        self.actionRecherche_Par_CIN_2.triggered.connect(self.affiche_loc_cin)
        self.actionRecherche_Par_Matricule_2.triggered.connect(self.affiche_loc_mat)
        self.actionRecherche_Par_Date_Donn_e.triggered.connect(self.affiche_loc_date)
        self.actionRecherche_Par_Dur_e_De_Location.triggered.connect(self.affiche_loc_duree)
        self.actionRecherche_Des_Locations_Entre_2_Dates_Donn_e.triggered.connect(self.affiche_loc_deux_date)
        self.actionEnregistrement_Fichiers_Voitures.triggered.connect(self.enregistre_voiture)
        self.actionR_cup_ration_fichier_Voitures.triggered.connect(self.recupere_voiture)
        self.actionEnregistrement_Fichier_Clients.triggered.connect(self.enregistre_client)
        self.actionR_cup_ration_Fichier_Clients.triggered.connect(self.recupere_client)
        self.actionEbregistrement_Fichier_Locations.triggered.connect(self.enregistre_location)
        self.actionR_cuperation_Fichier_Locations.triggered.connect(self.recupere_location)
        self.actionOuvrir.triggered.connect(self.enoncee)
        self.actionQuitter.triggered.connect(self.quitter_fen)
    
    def enoncee(self) :
        webbrowser.open_new('Projet Python Gestion du voiture.pdf')
    def quitter_fen(self):
        quit()
    def gotoaddv(self):
        ajoutvoiture = AjouterV()
        widget.addWidget(ajoutvoiture)
        widget.setCurrentWidget(ajoutvoiture)
    def suppr_par_matr(self):
        supprimevoiture = Supprimer_matrV()
        widget.addWidget(supprimevoiture) 
        widget.setCurrentWidget(supprimevoiture)
    def suppr_par_marque(self):
        supprime_marque = Supprime_par_marque()
        widget.addWidget(supprime_marque) 
        widget.setCurrentWidget(supprime_marque)
    def suppr_sup_10ans(self):
        s=supprime_date_10ans()
        widget.addWidget(s) 
        widget.setCurrentWidget(s)
    def modifie_prix(self):
        modifie = modifier_prix()
        widget.addWidget(modifie) 
        widget.setCurrentWidget(modifie)
    def modifie_couleur(self):
        modifie = modifier_couleur()
        widget.addWidget(modifie) 
        widget.setCurrentWidget(modifie)
    def affiche_dict_voiture(self):
        affiche=affichier_dict_voiture()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def affiche_par_matr(self):
        affiche = affiche_par_matricule()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def chercher_voit_couleur(self):
        affiche= cherchecolorv()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def chercher_voiture_marque(self):
        affiche=chercher_marque()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def chercher_voiture_disp(self):
        affiche=disponible()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def chercher_voiture_louee(self):
        affiche=louee()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def cherche_deux_date(self):
        affiche=rechercheentedeuxdates()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def ajout_client(self):
        ajout=ajouter_client()
        widget.addWidget(ajout) 
        widget.setCurrentWidget(ajout)
    def supp_client_cin(self):
        supp=Supprimer_cin()
        widget.addWidget(supp) 
        widget.setCurrentWidget(supp)
    def modifiadresse(self):
        modif=modifier_adresse()
        widget.addWidget(modif) 
        widget.setCurrentWidget(modif)
    def modifitelephone(self):
        modif=modifier_telephone()
        widget.addWidget(modif) 
        widget.setCurrentWidget(modif)
    def modifimail(self):
        modif=modifier_mail()
        widget.addWidget(modif) 
        widget.setCurrentWidget(modif)
    def affiche_dict_client(self):
        affiche =affichier_dict_client()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def affiche_cin(self):
        affiche= affiche_par_cin()
        widget.addWidget(affiche) 
        widget.setCurrentWidget(affiche)
    def ajout_location(self) :
        ajout=ajouter_location()
        widget.addWidget(ajout) 
        widget.setCurrentWidget(ajout)
    def supp_location(self) :
        s=supprimer_location()
        widget.addWidget(s) 
        widget.setCurrentWidget(s)
    def mod_date_locat(self):
        m=modifier_date_location()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def mod_duree_locat(self):
        m=modifier_duree_location()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def afficher_loc(self):
        m=affichier_dict_location()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def affiche_loc_cin(self):
        m=affichier_location_cin()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def affiche_loc_mat(self):
        m=affichier_location_mat()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def affiche_loc_date(self):
        m=affichier_location_date()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def affiche_loc_duree(self):
        m=affichier_location_duree()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def affiche_loc_deux_date(self):
        m=affiche_loc_deux_date()
        widget.addWidget(m) 
        widget.setCurrentWidget(m)
    def enregistre_voiture(self):
        wb=load_workbook('enregistrement de donnée.xlsx')
        ws=wb.active
        global voiture
        def enregistrement_fichierv(voiture):
            ws=wb['Voiture']
            i=1
            for key in voiture.keys():
                i=i+1
                A='A'+str(i)
                ws[A].value = key
                B='B'+str(i)
                ws[B].value = voiture[key][0]
                C='C'+str(i)
                ws[C].value = voiture[key][1]
                D='D'+str(i)
                ws[D].value = voiture[key][2]
                E='E'+str(i)
                ws[E].value = voiture[key][3]
                F='F'+str(i)
                ws[F].value = voiture[key][4]
            wb.save('enregistrement de donnée.xlsx')
        
        enregistrement_fichierv(voiture)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Enregistrement Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_() 
    def recupere_voiture(self) :
        wb=load_workbook('enregistrement de donnée.xlsx')
        ws=wb.active
        global voiture
        def recuperation_voitures(voiture):
            ws=wb['Voiture']
            L=[]
            row=2  
            while ws['A'+str(row)].value != None :
                for col in range(2,7):
                    char= get_column_letter(col)
                    cell= char + str(row)
                    L.append(ws[cell].value)
        
                voiture[ws['A'+str(row)].value]=L
                L=[]
                row=row+1
            L=[]
        recuperation_voitures(voiture)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Récupération Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_()
    def enregistre_client(self):
        global client
        def enregistre(client):
            wb=load_workbook('enregistrement de donnée.xlsx')
            ws=wb['Client']
            i=1
            for key in client.keys():
                i=i+1
                A='A'+str(i)
                ws[A].value = key
                B='B'+str(i)
                ws[B].value = client[key][0]
                C='C'+str(i)
                ws[C].value = client[key][1]
                D='D'+str(i)
                ws[D].value = client[key][2]
                E='E'+str(i)
                ws[E].value = client[key][3]
                F='F'+str(i)
                ws[F].value = client[key][4]
                G='G'+str(i)
                ws[G].value = client[key][5]
            wb.save('enregistrement de donnée.xlsx')
        enregistre(client)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Enregistrement Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_() 
    
    def recupere_client(self):
        global client
        wb=load_workbook('enregistrement de donnée.xlsx')
        ws=wb.active
        def recuperation_client(client):
            ws=wb['Client']
            L=[]
            row=2  
            while ws['A'+str(row)].value!=None :
                for col in range(2,8):
                    char= get_column_letter(col)
                    cell= char + str(row)
                    L.append(ws[cell].value)
        
                client[ws['A'+str(row)].value]=L
                L=[]
                row+=1
        recuperation_client(client)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Récupération Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_()
    def enregistre_location(self):
        global location     
        def enregistrement_fichierl(location):
            wb=load_workbook('enregistrement de donnée.xlsx')
            ws=wb['Location']
            
            i=1
            for key in location.keys():
                i=i+1
                A='A'+str(i)
                ws[A].value = key
                B='B'+str(i)
                ws[B].value = location[key][0]
                C='C'+str(i)
                ws[C].value = location[key][1]
                D='D'+str(i)
                ws[D].value = location[key][2]
                E='E'+str(i)
                ws[E].value = location[key][3]
                F='F'+str(i)
                ws[F].value = location[key][4]
                G='G' +str(i)
                ws[G].value = location[key][5]
            wb.save('enregistrement de donnée.xlsx')
        enregistrement_fichierl(location)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Enregistrement Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_() 
    def recupere_location(self):
        global location
        wb=load_workbook('enregistrement de donnée.xlsx')
        ws=wb.active
        def recuperation_location(location):
            ws=wb['Location']
            L=[]
            row=2  
            while ws['A'+str(row)].value!=None :
                for col in range(2,8):
                    char= get_column_letter(col)
                    cell= char + str(row)
                    if col ==4 :
                        x=ws[cell].value
                        y=datetime.date(x.year,x.month,x.day)
                        L.append(y)
                       
                    else :
                        L.append(ws[cell].value)
                    
                location[ws['A'+str(row)].value ]=L
                L=[]
                row+=1
        recuperation_location(location)
        msge = QtWidgets.QMessageBox()
        msge.setWindowTitle("Succée")
        msge.setText("Récupération Reussi !!")
        msge.setInformativeText("")
        msge.setIcon(QtWidgets.QMessageBox.Critical)
        x = msge.exec_()
        
        
#   ***** AJOUT DE VOITURE *****

class AjouterV(QDialog):
    def __init__(self):
        super(AjouterV,self).__init__()
        loadUi("ajouter_voiture.ui",self)
        
        self.retour.clicked.connect(self.quitter)
        self.AJOUTER.clicked.connect(self.enregistrer)
 
    
    
    def quitter(self):
        widget.setCurrentIndex(0)
        
    
    def enregistrer(self) :
        global voiture
        mat=self.lire_matricule.text()
        couleur=self.lire_couleur.text()
        marque=self.lire_marque.text()
        prix_location=self.lire_location.text()
        date=self.lire_date.date()
        ok=1
        
        
            
            
        if mat==""  :
            self.lire_matricule.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif verife_matr(mat,voiture) :
            self.lire_matricule.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("La Matricule Que Vous Avez Saisie Existe Deja !!")
            msge.setInformativeText("Redonner Un Autre Matricule")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif marque==""  :
            self.lire_marque.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Marque !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif couleur==""  :
            self.lire_couleur.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Couleur !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif couleur.isnumeric() :
            self.lire_couleur.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Couleur !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif prix_location==""  :
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Prix !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif (prix_location.isalpha()):
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Prix Doit Etre Des Chiffres Numerique !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif (len(prix_location)>1 and prix_location[0]=="-"):
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Erreur !! Le Prix Doit Etre Positive")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()                        
        
            
        elif self.boutton_disponible.isChecked():
            etat=self.boutton_disponible.text()            
        elif self.boutton_louee.isChecked():
            etat=self.boutton_louee.text()            
        else:
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer L'Etat Du Voiture !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        
        
        if(ok):
            ch=datetime.date(date.year(),date.month(),date.day())
            voiture[mat]=marque,couleur,etat,ch,prix_location
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succès")
            msge.setText("Ajout Reussi")
            msge.setInformativeText("Vouler Vous Ajouter Une Autre Voiture ?")
            msge.setIcon(QtWidgets.QMessageBox.Question)
            msge.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            msge.buttonClicked.connect(self.fate)
            x = msge.exec_()
        
        
    def fate(self,i):
        if i.text()=="&Yes":
            self.lire_matricule.clear()
            self.lire_couleur.clear()
            self.lire_marque.clear()
            self.lire_location.clear()
        else:
            widget.setCurrentIndex(0)




#  *****SUPPRIME DU VOITURE PAR MATRICULE *****



class Supprimer_matrV(QDialog):
    def __init__(self):
        super(Supprimer_matrV,self).__init__()
        loadUi("supprimer_par_matr.ui",self)
        self.retour.clicked.connect(self.quitter)
        self.supprimer.clicked.connect(self.kill)
    def quitter(self):
        widget.setCurrentIndex(0)
        
     
    def supp_par_matr(self,matr,d) :
        d1=dict()
        for i in d :
            if i != matr :
                d1[i]=d[i]
        return d1  
   
    def kill(self):
        global voiture
        mat=self.lire_matricule_supp.text()
        if voiture == {} :
            self.lire_matricule_supp.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif mat==""  :
            self.lire_matricule_supp.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif verife_matr(mat,voiture) :
            self.lire_matricule_supp.clear()
            voiture=self.supp_par_matr(mat,voiture)
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succès")
            msge.setText("La Voiture A Ete Supprimer Avec Succée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        else :
            self.lire_matricule_supp.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("Cette Voiture N'existe Pas !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        
#*******SUPPRIME DU VOITURE PAR MARQUE *****
        
    
class Supprime_par_marque(QDialog):
    def __init__(self):
        super(Supprime_par_marque,self).__init__()
        loadUi("supprimer_par_marque.ui",self)
        self.retour.clicked.connect(self.quitter)
        self.supprimer.clicked.connect(self.kill)
    def quitter(self):
        widget.setCurrentIndex(0)
     
    
    def supp_par_marque(self,marque,d) :
        d1=dict()
        ok=0
        for i in d :
            l=[]
            l=d[i]
            if marque not in l :
                d1[i]=d[i]
        if d!=d1 :
            ok =1
        return d1,ok
    
    def kill(self):
        global voiture
        marque=self.lire_marque.text()
        if voiture == {} :
            self.lire_marque.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif marque==""  :
            self.lire_marque.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Marque !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            voiture,ok=self.supp_par_marque(marque,voiture)
            if ok :
                self.lire_marque.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("Les Voitures Sont Etes Supprimer Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
                print(voiture)
            else :
                self.lire_marque.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Echec")
                msge.setText("Ils N'existe Aucune Voiture De cette Marque !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()


#***** supprimer des voitures age >10  ans *****
                
            
class supprime_date_10ans(QDialog):
    def __init__(self):
        super(supprime_date_10ans,self).__init__()
        loadUi("supprime_voiture_par_age.ui",self)
        self.retour.clicked.connect(self.Return)
        self.supprimer.clicked.connect(self.supp)    
    def Return(self):
        widget.setCurrentIndex(0)
    def supp(self):
        global voiture
        d=dict()
        dt2=datetime.date(2010,1,1)-datetime.date(2000,1,4)
        dt3=datetime.date(datetime.datetime.now().year,datetime.datetime.now().month,datetime.datetime.now().day)
        ok=1
        for key in voiture.keys():
            jaar=int(voiture[key][3].year)
            monat=int(voiture[key][3].month)
            tag=int(voiture[key][3].day)
            dt1=datetime.date(jaar,monat,tag) 
            if dt3-dt1 <= dt2 :
                d[key]=voiture[key]
            else :
                ok=0
        voiture=d
        if voiture=={} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif ok == 0 :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succée")
            msge.setText("Suppression Reussi!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        else :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Désolée")
            msge.setText("Aucune Voiture D'age Supérieure A 10 Ans!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)



#*****MODIFIER LE PRIX DE VOITURE ******
                
class modifier_prix(QDialog):
    def __init__(self):
        super(modifier_prix,self).__init__()
        loadUi("modifier_prix.ui",self)
        self.annuler.clicked.connect(self.quitter)
        self.modifier.clicked.connect(self.modif)
    def quitter(self):
        widget.setCurrentIndex(0)
    
    
    def modifierprix(self,matr,d,prix) :
        for i in d :
            if i == matr :
                l=list(d[i])
                l.pop(4)
                l.append(prix)
                l=tuple(l) 
                d[i]=l
    
    def modif(self):
        mat=self.lire_mat.text()
        
        if voiture == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!Essayer De Le Remplir")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif mat==""  :
            self.lire_mat.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif verife_matr(mat,voiture) :
            prix=self.lire_prix.text()
            if prix==""  :
                self.lire_prix.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Prix !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif (prix.isalpha()):
                self.lire_prix.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Le Prix Doit Etre Des Chiffres Numerique !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif (len(prix)>1 and prix[0]=="-"):
                self.lire_prix.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Erreur !! Le Prix Doit Etre Positive")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                self.lire_prix.clear()
                self.lire_mat.clear()
                self.modifierprix(mat,voiture,prix)
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("Le Prix Du Voiture A Ete Modifier Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
                widget.setCurrentIndex(0)
            
        else :
            self.lire_mat.clear()
            self.lire_prix.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("Cette Voiture N'existe Pas !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()


            
#*****MODIFIER LE COULEUR DE VOITURE ******
                
class modifier_couleur(QDialog):
    def __init__(self):
        super(modifier_couleur,self).__init__()
        loadUi("modifier_couleur.ui",self)
        self.annuler.clicked.connect(self.quitter)
        self.modifier.clicked.connect(self.modif)
    def quitter(self):
        widget.setCurrentIndex(0)
    
    
    def modifiercouleur(self,matr,d,couleur) : 
        for i in d :
            if i == matr :
                l=list(d[i])
                l.pop(1)
                l.insert(1,couleur)
                l=tuple(l) 
                d[i]=l

    
    def modif(self):
        mat=self.lire_mat.text()
        
        if voiture == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!Essayer De Le Remplir")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif mat==""  :
            self.lire_mat.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif verife_matr(mat,voiture) :
            couleur=self.lire_couleur.text()
            if couleur==""  :
                self.lire_couleur.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Couleur !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif couleur.isnumeric():
                self.lire_couleur.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Le Couleur Doit Etre Composée Par Des Lettres Alphabéthique !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                self.lire_couleur.clear()
                self.lire_mat.clear()
                self.modifiercouleur(mat,voiture,couleur)
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("Le Couleur Du Voiture A Ete Modifier Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
                widget.setCurrentIndex(0)
        else :
            self.lire_mat.clear()
            self.lire_couleur.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("IL N'existe Aucun Voiture Avec Cette Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()


#*****AFFICHER LE CONTENUE DU DICTIONNAIRE DU VOITUTRE *****
            
class affichier_dict_voiture(QDialog):
    def __init__(self):
        super(affichier_dict_voiture,self).__init__()
        loadUi("tab_dict_voiture.ui",self)
        self.loaddata()
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global voiture
        row=0
        i=1
        self.tableWidget.setRowCount(len(voiture))
        for key in voiture.keys():
            date=voiture[key][3]
            a=date.year
            m=date.month
            j=date.day
            ch=str(j)+"-"+str(m)+"-"+str(a)
            
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"voiture N°{i}"))
            self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(key))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[key][0]))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[key][1]))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[key][2]))            
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[key][4])))            
            row=row+1
            i=i+1



#**** AFFICHE DU VOITURE PAR MATRICULE *****
            
            
            
            
class affiche_par_matricule(QDialog):
    def __init__(self):
        super(affiche_par_matricule,self).__init__()
        loadUi("affiche_par_matr.ui",self)
        self.afficher.clicked.connect(self.afficher_voiture)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
    def afficher_voiture(self):
        mat=self.lire_matricule_aff.text()
        if voiture =={} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif mat==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif verife_matr(mat,voiture)==0 :
            self.lire_matricule_aff.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Désolée")
            msge.setText("Matricule Inexistant !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            for i in voiture :
                if i==mat :
                    self.tableWidget.setRowCount(1)
                    date=voiture[i][3]
                    a=date.year
                    m=date.month
                    j=date.day
                    ch=str(j)+"-"+str(m)+"-"+str(a)
                    
                    self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem(i))
                    self.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(voiture[i][0]))
                    self.tableWidget.setItem(0, 2, QtWidgets.QTableWidgetItem(voiture[i][1]))
                    self.tableWidget.setItem(0, 3, QtWidgets.QTableWidgetItem(voiture[i][2]))            
                    self.tableWidget.setItem(0, 4, QtWidgets.QTableWidgetItem(ch))            
                    self.tableWidget.setItem(0, 5, QtWidgets.QTableWidgetItem(str(voiture[i][4])))            
                    
                    

#*****recherche de voiture par couleur***
class cherchecolorv(QDialog):
    def __init__(self):
        super(cherchecolorv,self).__init__()
        loadUi("affiche_par_couleur.ui",self)
        self.retour.clicked.connect(self.Return)
        self.afficher.clicked.connect(self.find)     
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global voiture
        L=[]
        ok=0
        col=self.lire_couleur.text()
        if voiture =={} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0) 
        elif col=="" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Couleur !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else:
            for key in voiture.keys():
                if voiture[key][1]==col:
                    L.append(key)
                    ok=1
            self.tableWidget.setRowCount(len(L))
            row=0
            f=1
            for i in L:
                date=voiture[i][3]
                a=date.year
                m=date.month
                j=date.day
                ch=str(j)+"-"+str(m)+"-"+str(a)
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Voiture {f} "))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(i))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[i][0]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[i][1]))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[i][2]))            
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
                self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[i][4])))
                row=row+1
                f=f+1
            if ok == 0 :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Voiture Avec Cette couleur !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
         
#*****RECHERCHE DE VOITURE PAR MARQUE *****
                
class chercher_marque(QDialog):
    def __init__(self):
        super(chercher_marque,self).__init__()
        loadUi("affiche_par_marque.ui",self)
        self.retour.clicked.connect(self.Return)
        self.afficher.clicked.connect(self.find)     
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global voiture
        L=[]
        ok=0
        col=self.lire_couleur.text()
        if voiture == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif col=="" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Marque !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else:
            for key in voiture.keys():
                if voiture[key][0]==col:
                    L.append(key)
                    ok=1
            self.tableWidget.setRowCount(len(L))
            row=0
            f=1
            for i in L:
                date=voiture[i][3]
                a=date.year
                m=date.month
                j=date.day
                ch=str(j)+"-"+str(m)+"-"+str(a)
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Voiture {f} "))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(i))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[i][0]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[i][1]))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[i][2]))            
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
                self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[i][4])))
                row=row+1
                f=f+1
            if ok == 0 :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Voiture Avec Cette Marque !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
         
#**** recherche des voitures disponibles ****


class disponible(QDialog):
    def __init__(self):
        super(disponible,self).__init__()
        loadUi("affiche_disponible.ui",self)
        self.retour.clicked.connect(self.Return)
        self.find()     
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global voiture
        L=[]
        for key in voiture.keys():
            if voiture[key][2]=="Disponible":
                L.append(key)
        self.tableWidget.setRowCount(len(L))
        row=0
        f=1
        for i in L:
            date=voiture[i][3]
            a=date.year
            m=date.month
            j=date.day
            ch=str(a)+"-"+str(m)+"-"+str(j)
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Voiture {f}"))
            self.tableWidget.setItem(row, 1 , QtWidgets.QTableWidgetItem(i))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[i][0]))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[i][1]))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[i][2]))            
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[i][4])))            
            row=row+1
            f=f+1
        
        
#***** recherche des voitures louee *****
            
            
            
class louee(QDialog):
    def __init__(self):
        super(louee,self).__init__()
        loadUi("affiche_louee.ui",self)
        self.retour.clicked.connect(self.Return)
        self.find()     
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global voiture
        L=[]
        for key in voiture.keys():
            if voiture[key][2]=="Louée":
                L.append(key)
        self.tableWidget.setRowCount(len(L))
        row=0
        f=1
        for i in L:
            date=voiture[i][3]
            a=date.year
            m=date.month
            j=date.day
            ch=str(a)+"-"+str(m)+"-"+str(j)
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Voiture {f}"))
            self.tableWidget.setItem(row, 1 , QtWidgets.QTableWidgetItem(i))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[i][0]))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[i][1]))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[i][2]))            
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[i][4])))            
            row=row+1
            f=f+1
        
        
#**** recherche des voitures louee entres deux dates ****
            
class rechercheentedeuxdates(QDialog):
    def __init__(self):
        super(rechercheentedeuxdates,self).__init__()
        loadUi("affiche_entre_deux_date.ui",self)
        self.retour.clicked.connect(self.Return)
        self.afficher.clicked.connect(self.find)    
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global voiture
        L=[]
        ok=0
        date=self.dateEdit.date()
        date1=self.dateEdit_2.date()
        dt=datetime.timedelta(1)
        a=date.year()
        m=date.month()
        j=date.day()
        datep1=datetime.date(a,m,j)

        a1=date1.year()
        m1=date1.month()
        j1=date1.day()
        datep2=datetime.date(a1,m1,j1)
        if voiture == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Voiture Est Vide !!!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif datep1>datep2:
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("La Premier Date Doit Etre Inferieur A La Deuxiéme !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        else :
            for key in voiture.keys():
                jaar=int(voiture[key][3].year)
                monat=int(voiture[key][3].month)
                tag=int(voiture[key][3].day)
                hayt=datetime.date(jaar,monat,tag) 
                if hayt>=datep1 and hayt<=datep2 and voiture[key][2]=="Louée":
                    L.append(key)
                    ok=1
            
        
            self.tableWidget.setRowCount(len(L))
            row=0
            f=1
            for i in L:
                date=voiture[i][3]
                a=date.year
                m=date.month
                j=date.day
                ch=str(j)+"-"+str(m)+"-"+str(a)
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Voiture N°{f}" ))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(i))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(voiture[i][0]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(voiture[i][1]))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(voiture[i][2]))            
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch))            
                self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(voiture[i][4])))            
                row=row+1
                f=f+1
            if ok == 0 :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Voiture Trouvée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()

#*****  GESTION DES CLIENTS ******#
                
                
#*** ajouter un nouvel clients *****
class ajouter_client(QDialog):
    def __init__(self):
        super(ajouter_client,self).__init__()
        loadUi("ajouter_client.ui",self)
        self.retour.clicked.connect(self.Return)
        self.ajouter.clicked.connect(self.add)           
    def Return(self):
        widget.setCurrentIndex(0)
        
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    
    def add(self):    
        global client
        ok=1
        cin=self.lire_cin.text()
        nom=self.lire_nom.text()
        prenom=self.lire_prenom.text()
        mail=self.lire_mail.text()
        age=self.lire_age.text()
        adresse=self.lire_adresse.text()
        telephone=self.lire_telephone.text()
         
        if cin==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif self.verife_cin(cin,client) :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("La CIN Que Vous Avez Saisie Existe Deja !!")
            msge.setInformativeText("Redonner Un Autre CIN")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif nom==""  :
            self.lire_nom.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Nom !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif nom.isnumeric():
            self.lire_nom.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Nom Invalide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif prenom==""  :
            self.lire_prenom.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Prenom !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif prenom.isnumeric():
            self.lire_prenom.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Prenom Invalide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()   
        
        elif mail==""  :
            self.lire_mail.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Mail !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif age=="" :
            self.lire_age.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer L'age Du Client !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        elif age.isalpha() :
            self.lire_age.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Age Invalide  !! Le Redonner !")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
         
        elif (len(age)>1 and age[0]=="-"):
            self.lire_age.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Erreur !! L'age Doit Etre Positive")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()                        
        
        elif adresse==""  :
            self.lire_adresse.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer L'adresse Du Client !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        
        elif telephone=="":
            self.lire_telephone.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Numéro De Telephone !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        
        elif not(telephone.isnumeric()) or len(telephone)!=8 :
            self.lire_telephone.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Numéro De Telephone Doit Etre Des 8 Chiffres Numerique !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            ok=0
            x = msge.exec_()
        
        
        if(ok):
            client[cin]=nom,prenom,mail,age,adresse,telephone
            print(client)
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succès")
            msge.setText("La Client A Eté Ajouter Avec Succée !!")
            msge.setInformativeText("Vouler Vous Ajouter Un Autre Client ?")
            msge.setIcon(QtWidgets.QMessageBox.Question)
            msge.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            msge.buttonClicked.connect(self.fate)
            x = msge.exec_()
        
        
    def fate(self,i):
        if i.text()=="&Yes":
            self.lire_cin.clear()
            self.lire_nom.clear()
            self.lire_prenom.clear()
            self.lire_mail.clear()
            self.lire_age.clear()
            self.lire_adresse.clear()
            self.lire_telephone.clear()
        else:
            widget.setCurrentIndex(0)

#***** supprimer un client par cin ****


class Supprimer_cin(QDialog):
    def __init__(self):
        super(Supprimer_cin,self).__init__()
        loadUi("supprimer_client_par_cin.ui",self)
        self.retour.clicked.connect(self.quitter)
        self.supprimer.clicked.connect(self.kill)
    def quitter(self):
        widget.setCurrentIndex(0)
     
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
     
    def supp_par_cin(self,cin,d) :
        d1=dict()
        for i in d :
            if i != cin :
                d1[i]=d[i]
        return d1  
   
    def kill(self):
        global client
        cin=self.lire_cin.text()
        if client == {} :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Client Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif cin==""  :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()    
            
        elif self.verife_cin(cin,client) :
            self.lire_cin.clear()
            client=self.supp_par_cin(cin,client)
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succès")
            msge.setText("Le Client A Eté Supprimer Avec Succée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        else :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("Ce Client N'existe Pas Dans Le Dictionnaire !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            

            
#*****MODIFIER L'ADRESSE DE CLIENT ******
                
class modifier_adresse(QDialog):
    def __init__(self):
        super(modifier_adresse,self).__init__()
        loadUi("modifier_adresse.ui",self)
        self.annuler.clicked.connect(self.quitter)
        self.modifier.clicked.connect(self.modif)
        
    def quitter(self):
        widget.setCurrentIndex(0)
    
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    
    def modifier_adresse(self,cin,adresse,d) :
        for i in d :
            if i == cin :
                l=list(d[i])
                l.pop(4)
                l.insert(4,adresse)
                l=tuple(l) 
                d[i]=l

    
    def modif(self):
        cin=self.lire_cin.text()
        
        if client == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Client Est Vide !!Essayer De Le Remplir")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif cin==""  :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()    
        
        elif self.verife_cin(cin,client) :
            adresse=self.lire_adresse.text()
            if adresse==""  :
                self.lire_adresse.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Nouveau Adresse Du Client  !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                self.lire_adresse.clear()
                self.lire_cin.clear()
                self.modifier_adresse(cin,adresse,client)
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("L'adresse Du Client A Eté Modifier Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
                
            
        else :
            self.lire_cin.clear()
            self.lire_adresse.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("IL N'existe Aucun Client Avec Cette N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()

        

#****modifier le telephonne du client ****
                
class modifier_telephone(QDialog):
    def __init__(self):
        super(modifier_telephone,self).__init__()
        loadUi("modifier_telephone.ui",self)
        self.annuler.clicked.connect(self.quitter)
        self.modifier.clicked.connect(self.modif)
        
    def quitter(self):
        widget.setCurrentIndex(0)
    
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    
    def modifier_telephone(self,cin,numero,d) :
        for i in d :
            if i == cin :
                l=list(d[i])
                l.pop(5)
                l.append(numero)
                l=tuple(l) 
                d[i]=l

    
    def modif(self):
        cin=self.lire_cin.text()
        
        if client == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Client Est Vide !!Essayer De Le Remplir")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif cin==""  :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()    
        
        elif self.verife_cin(cin,client) :
            telephone=self.lire_telephone.text()
            if telephone==""  :
                self.lire_telephone.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Nouveau N° De Telephone Du Client  !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif not(telephone.isnumeric()) or len(telephone)!=8 :
                self.lire_telephone.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Le Numéro De Telephone Doit Etre Des 8 Chiffres Numerique !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                self.lire_telephone.clear()
                self.lire_cin.clear()
                self.modifier_telephone(cin,telephone,client)
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("Le N° Du Telephone Du Client A Eté Modifier Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
             
            
        else :
            self.lire_cin.clear()
            self.lire_telephone.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("IL N'existe Aucun Client Avec Cette N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
           

            
#*****MODIFIER E_MAIL DE CLIENT ******
                
class modifier_mail(QDialog):
    def __init__(self):
        super(modifier_mail,self).__init__()
        loadUi("modifier_mail.ui",self)
        self.annuler.clicked.connect(self.quitter)
        self.modifier.clicked.connect(self.modif)
        
    def quitter(self):
        widget.setCurrentIndex(0)
    
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    
    
    def modifier_mail(self,cin,mail,d) :
        for i in d :
            if i == cin :
                l=list(d[i])
                l.pop(2)
                l.insert(2,mail)
                l=tuple(l) 
                d[i]=l


    
    def modif(self):
        cin=self.lire_cin.text()
        
        if client == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Le Dictionnaire Du Client Est Vide !!Essayer De Le Remplir")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
       
        elif cin==""  :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()    
        
        elif self.verife_cin(cin,client) :
            mail=self.lire_mail.text()
            if mail==""  :
                self.lire_mail.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Nouveau Adresse Mail Du Client  !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                self.lire_mail.clear()
                self.lire_cin.clear()
                self.modifier_mail(cin,mail,client)
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Succès")
                msge.setText("L'adresse Du Client A Eté Modifier Avec Succée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
               
            
        else :
            self.lire_cin.clear()
            self.lire_mail.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Echec")
            msge.setText("IL N'existe Aucun Client Avec Cette N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()


#**** afficher le contenu du dictionnaire de client *****
            
class affichier_dict_client(QDialog):
    def __init__(self):
        super(affichier_dict_client,self).__init__()
        loadUi("tab_dict_client.ui",self)
        self.loaddata()
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global client
        row=0
        i=1
        self.tableWidget.setRowCount(len(client))
        for key in client.keys():
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"Client N°{i}"))
            self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(key))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(client[key][0]))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(client[key][1]))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(client[key][2]))            
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(client[key][3]))            
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(client[key][4]))
            self.tableWidget.setItem(row, 7, QtWidgets.QTableWidgetItem(client[key][5]))
            row=row+1
            i=i+1
#**** affiche du client par cin *****
            
           
class affiche_par_cin(QDialog):
    def __init__(self):
        super(affiche_par_cin,self).__init__()
        loadUi("affiche_par_cin.ui",self)
        #self.loaddata()
        self.afficher.clicked.connect(self.afficher_client)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def verife_cin(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    def afficher_client(self):
        cin=self.lire_cin.text()
        if client == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Client Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif cin==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()    
        elif self.verife_cin(cin,client)==0 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Désolée")
            msge.setText("N° De CIN Inexistant !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            for i in client :
                if i==cin :
                    self.tableWidget.setRowCount(1)
                    self.tableWidget.setItem(0, 0, QtWidgets.QTableWidgetItem(i))
                    self.tableWidget.setItem(0, 1, QtWidgets.QTableWidgetItem(client[i][0]))
                    self.tableWidget.setItem(0, 2, QtWidgets.QTableWidgetItem(client[i][1]))
                    self.tableWidget.setItem(0, 3, QtWidgets.QTableWidgetItem(client[i][2]))            
                    self.tableWidget.setItem(0, 4, QtWidgets.QTableWidgetItem(client[i][3]))            
                    self.tableWidget.setItem(0, 5, QtWidgets.QTableWidgetItem(client[i][4]))            
                    self.tableWidget.setItem(0, 6, QtWidgets.QTableWidgetItem(client[i][5]))



#**** GESTION DE LOACTION *****

#**** ajouter location ****
           
class ajouter_location(QDialog):
    def __init__(self):
        super(ajouter_location,self).__init__()
        loadUi("ajouter_location.ui",self)
        self.ajouter.clicked.connect(self.ajoutervoit)
        self.retour.clicked.connect(self.Return)
        
    def cherche_disponible(self,d):
        for i in d :
            if d[i][2] == "Disponible" :
                return 1
                break
        return 0
    
    def Return(self):
        widget.setCurrentIndex(0)
        
    def client_existe(self,cin,d) :
        for i in d :
            if i==cin :
                return 1
                break
        return 0
    
    def voiture_existe(self,mat,d):
        for i in d :
            if i==mat :
                return 1
                break
        return 0
    
    def fate(self,i):
        if i.text()=="&Yes":
            self.lire_matr.clear()
            self.lire_location.clear()
            self.lire_cin.clear()
            self.lire_duree.clear()
        else:
            widget.setCurrentIndex(0)

    def verife_num(self,num,d) :
        for i in d :
            if i==num :
                return 1
                break
        return 0
    
    def voiture_dispo(self,mat,d) :
        for i in d :
            if i == mat :
                if d[i][2]=="Disponible" :
                    return 1
                    break
        return 0
    
    def ajoutervoit(self) :
        global voiture
        global location
        num_location=self.lire_location.text()
        cin=self.lire_cin.text()
        mat=self.lire_matr.text()
        date=self.dateEdit.date()
        heure=self.timeEdit.time()
        duree=self.lire_duree.text()
        
        if self.cherche_disponible(voiture) :
            if num_location == "" :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Erreur")
                msge.setText("Faire Indiquer Le Numéro Du Location !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif not(num_location.isnumeric()):
                self.lire_location.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Erreur")
                msge.setText("Numéro Du Location Doit Etre Des Chiffres Numérique !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif self.verife_num(num_location,location) :
                self.lire_location.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Erreur")
                msge.setText("Numéro Du Location Existe Dejà !!")
                msge.setInformativeText(" Ressayer Avec Un Autre !")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            
            elif cin==""  :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le N° De CIN !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            
            elif not(cin.isnumeric()) or len(cin)!=8 :
                self.lire_cin.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif mat =="" :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer Le Matricule !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif duree == "" :
               
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Faire Indiquer La Durée Du Location !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif not(duree.isnumeric()) :
                self.lire_duree.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Durée Du Location Doit Etre Des Chiffres Numérique!!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            elif self.client_existe(cin,client) :
                if self.voiture_existe(mat,voiture) :
                    if self.voiture_dispo(mat,voiture) :
                        ch=datetime.date(date.year(),date.month(),date.day())
                        ch1=datetime.time(heure.hour(),heure.minute())
                        montant=str(float(voiture[mat][4])*int(duree))
                        location[num_location]=cin,mat,ch,ch1,duree,montant
                        l=[]
                        l=list(voiture[mat])
                        l.pop(2)
                        l.insert(2,"Louée")
                        voiture[mat]=tuple(l)
                        
                        msge = QtWidgets.QMessageBox()
                        msge.setWindowTitle("Succès")
                        msge.setText("La Location A Eté Enrregistrer Avec Succée ")
                        msge.setInformativeText("Vouler Vous Ajouter Une Autre Location ?")
                        msge.setIcon(QtWidgets.QMessageBox.Question)
                        msge.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                        msge.buttonClicked.connect(self.fate)
                        x = msge.exec_()
                        
                    else :
                        self.lire_location.clear()
                        self.lire_cin.clear()
                        self.lire_matr.clear()
                        self.lire_duree.clear()
                        msge = QtWidgets.QMessageBox()
                        msge.setWindowTitle("Désolée")
                        msge.setText("La Voiture Voulu Est Louée !!")
                        msge.setInformativeText("")
                        msge.setIcon(QtWidgets.QMessageBox.Critical)
                        x = msge.exec_() 
                    
                else :
                    self.lire_location.clear()
                    self.lire_cin.clear()
                    self.lire_matr.clear()
                    self.lire_duree.clear()
                    msge = QtWidgets.QMessageBox()
                    msge.setWindowTitle("Attention")
                    msge.setText("Cette Voiture N'est Pas Declarer Dans Le Dictionnaire Du Voiture !!")
                    msge.setInformativeText("Essayer De L'enrregistrer Puis Ressayer !!")
                    msge.setIcon(QtWidgets.QMessageBox.Critical)
                    x = msge.exec_()
                
                
            else :
                self.lire_location.clear()
                self.lire_cin.clear()
                self.lire_matr.clear()
                self.lire_duree.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Attention")
                msge.setText("Cette Personne N'est Pas Declarer Dans Le Dictionnaire Du Client !!")
                msge.setInformativeText("Essayer De L'enrregistrer Puis Ressayer !!")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
                
                
               
           
           
        else :
            self.lire_location.clear()
            self.lire_cin.clear()
            self.lire_matr.clear()
            self.lire_duree.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Désolée")
            msge.setText("IL N'existe Aucune Voiture Disponible !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
            
#***** supprimer une location ******
class supprimer_location(QDialog):
    def __init__(self):
        super(supprimer_location,self).__init__()
        loadUi("supprime_location.ui",self)
        self.supprimer.clicked.connect(self.supprimer_loc)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
     
    def verife_num(self,num,d) :
        for i in d :
            if i==num :
                return 1
                break
        return 0
    def supp(self,num,d) :
        d1=dict()
        for i in d :
            if i!=num :
                d1[i]=d[i]
        return d1
        
    
    def supprimer_loc(self):
        global location
        global voiture
        num_location=self.lire_location.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif num_location == "" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Faire Indiquer Le Numéro Du Location !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif not(num_location.isnumeric()):
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location Doit Etre Des Chiffres Numérique !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif self.verife_num(num_location,location)==0 :
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location N'Existe Pas Dejà !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        
        else :
            self.lire_location.clear()
            mat=location[num_location][1]
            location=self.supp(num_location,location)
            l=[]
            l=list(voiture[mat])
            l.pop(2)
            l.insert(2,"Disponible")
            voiture[mat]=tuple(l)
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succée")
            msge.setText(" Location Supprimé Avec Succée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            


# **** modifier la date d'une location *****


class modifier_date_location(QDialog):
    def __init__(self):
        super(modifier_date_location,self).__init__()
        loadUi("modifier_date_location.ui",self)
        self.modifier.clicked.connect(self.modifie_loc)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
     
    def verife_num(self,num,d) :
        for i in d :
            if i==num :
                return 1
                break
        return 0
    def modifie_loc(self) :
        global location
        num=self.lire_location.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif num == "" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Faire Indiquer Le Numéro Du Location !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif not(num.isnumeric()):
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location Doit Etre Des Chiffres Numérique !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif self.verife_num(num,location)==0 :
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location N'Existe Pas !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            date=self.dateEdit.date()
            heure=self.timeEdit.time()
            ch=datetime.date(date.year(),date.month(),date.day())
            ch1=datetime.time(heure.hour(),heure.minute())
            l=[]
            l=list(location[num])
            l.pop(2)
            l.insert(2,ch)
            l.pop(3)
            l.insert(3,ch1)
            location[num]=tuple(l)
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succée")
            msge.setText("Modification Reussi !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
           
#*** modifier duréee d'une location ****
            
            
class modifier_duree_location(QDialog):
    def __init__(self):
        super(modifier_duree_location,self).__init__()
        loadUi("modifier_duree_location.ui",self)
        self.modifier.clicked.connect(self.modifie_loc)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
     
    def verife_num(self,num,d) :
        for i in d :
            if i==num :
                return 1
                break
        return 0
    def modifie_loc(self) :
        global location
        global voiture
        num=self.lire_location.text()
        duree=self.lire_duree.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif num == "" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Faire Indiquer Le Numéro Du Location !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif not(num.isnumeric()):
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location Doit Etre Des Chiffres Numérique !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif self.verife_num(num,location)==0 :
            self.lire_location.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Numéro Du Location N'Existe Pas !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif duree == "" :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("Veuillez Donner Le Nouveau Durée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif not(duree.isnumeric()) :
            self.lire_duree.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Durée Du Location Doit Etre Des Chiffres Numérique!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            mat=location[num][1]
            montant=str(int(voiture[mat][4])*int(duree))
            l=[]
            l=list(location[num])
            l.pop(4)
            l.insert(4,duree)
            l.pop(5)
            l.append(montant)
            location[num]=tuple(l)
            self.lire_location.clear()
            self.lire_duree.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Succée")
            msge.setText("Modification Reussi !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
           


#**** contenue de dictionnaire du location *****
            
class affichier_dict_location(QDialog):
    def __init__(self):
        super(affichier_dict_location,self).__init__()
        loadUi("tab_dict_location.ui",self)
        self.loaddata()
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global location
        row=0
        i=1
        self.tableWidget.setRowCount(len(location))
        for key in location.keys():
            date=location[key][2]
            time=location[key][3]
            a=date.year
            m=date.month
            j=date.day
            h=time.hour
            m1=time.minute
            ch=str(j)+"-"+str(m)+"-"+str(a)
            ch1=str(h)+":"+str(m1)
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(f"location N°{i}"))
            self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(key))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[key][0]))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(location[key][1]))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch))
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(ch1))
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[key][4]))            
            self.tableWidget.setItem(row, 7, QtWidgets.QTableWidgetItem(location[key][5]))            
            row=row+1
            i=i+1


# *** affiche location par cin *****

            
class affichier_location_cin(QDialog):
    def __init__(self):
        super(affichier_location_cin,self).__init__()
        loadUi("affiche_par_cin_location.ui",self)
        self.afficher.clicked.connect(self.loaddata)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global location
        cin=self.lire_cin.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif cin==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le N° De CIN !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        elif not(cin.isnumeric()) or len(cin)!=8 :
            self.lire_cin.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText(" CIN Invalide !! CIN Doit Etre Composé Des 8 Chiffres ")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        else :
            ok=0
            l=[]
            for i in location :
                if location[i][0]==cin :
                    l.append(i)
                    ok=1
            if ok :
                row=0
                self.tableWidget.setRowCount(len(l))
                for i in l :
                    date=location[i][2]
                    time=location[i][3]
                    a=date.year
                    m=date.month
                    j=date.day
                    h=time.hour
                    m1=time.minute
                    ch=str(j)+"-"+str(m)+"-"+str(a)
                    ch1=str(h)+":"+str(m1)
                    self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(i))
                    self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(location[i][0]))
                    self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[i][1]))
                    self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(ch))
                    self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch1))
                    self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(location[i][4]))            
                    self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[i][5]))            
                    row=row+1
            else :
                self.lire_cin.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Location Correspondant A Cette Personne !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()  



# *** affiche location par matricule *****

            
class affichier_location_mat(QDialog):
    def __init__(self):
        super(affichier_location_mat,self).__init__()
        loadUi("affiche_par_mat_location.ui",self)
        self.afficher.clicked.connect(self.loaddata)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global location
        mat=self.lire_mat.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif mat==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer Le Matricule !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            
        else :
            ok=0
            l=[]
            for i in location :
                if location[i][1]==mat :
                    l.append(i)
                    ok=1
            if ok :
                row=0
                self.tableWidget.setRowCount(len(l))
                for i in l :
                    date=location[i][2]
                    time=location[i][3]
                    a=date.year
                    m=date.month
                    j=date.day
                    h=time.hour
                    m1=time.minute
                    ch=str(j)+"-"+str(m)+"-"+str(a)
                    ch1=str(h)+":"+str(m1)
                    self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(i))
                    self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(location[i][0]))
                    self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[i][1]))
                    self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(ch))
                    self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch1))
                    self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(location[i][4]))            
                    self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[i][5]))            
                    row=row+1
            else :
                self.lire_mat.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Location Correspondant A Cette Voiture !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()  

#*****recherche de location par date donnée *****
            
class affichier_location_date(QDialog):
    def __init__(self):
        super(affichier_location_date,self).__init__()
        loadUi("affiche_location_date.ui",self)
        self.afficher.clicked.connect(self.loaddata)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
    def loaddata(self) :
        date=self.dateEdit.date()
        time=self.timeEdit.time()
        ch=datetime.date(date.year(),date.month(),date.day())
        ch1=datetime.time(time.hour(),time.minute())
        l=[]
        ok=0
        for i in location :
            if location[i][2]==ch and location[i][3]==ch1 :
                l.append(i)
                ok=1
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif ok :
            row=0
            self.tableWidget.setRowCount(len(l))
            for i in l :
                date=location[i][2]
                time=location[i][3]
                a=date.year
                m=date.month
                j=date.day
                h=time.hour
                m1=time.minute
                ch=str(j)+"-"+str(m)+"-"+str(a)
                ch1=str(h)+":"+str(m1)
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(i))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(location[i][0]))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[i][1]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(ch))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch1))
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(location[i][4]))            
                self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[i][5]))            
                row=row+1
        else :
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Location Correspondant A Cette Date !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()  
        
#**** recherche des locations par durée de location ****
                           
class affichier_location_duree(QDialog):
    def __init__(self):
        super(affichier_location_duree,self).__init__()
        loadUi("affiche_par_duree_location.ui",self)
        self.afficher.clicked.connect(self.loaddata)
        self.retour.clicked.connect(self.Return)
        
    def Return(self):
        widget.setCurrentIndex(0)
        
    def loaddata(self):
        global location
        duree=self.lire_duree.text()
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif duree==""  :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Faire Indiquer La Durée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif not(duree.isnumeric()) :
            self.lire_duree.clear()
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Durée Du Location Doit Etre Des Chiffres Numérique!!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()  
        else :
            ok=0
            l=[]
            for i in location :
                if location[i][4]==duree :
                    l.append(i)
                    ok=1
            if ok :
                row=0
                self.tableWidget.setRowCount(len(l))
                for i in l :
                    date=location[i][2]
                    time=location[i][3]
                    a=date.year
                    m=date.month
                    j=date.day
                    h=time.hour
                    m1=time.minute
                    ch=str(j)+"-"+str(m)+"-"+str(a)
                    ch1=str(h)+":"+str(m1)
                    self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(i))
                    self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(location[i][0]))
                    self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[i][1]))
                    self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(ch))
                    self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch1))
                    self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(location[i][4]))            
                    self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[i][5]))            
                    row=row+1
            else :
                self.lire_duree.clear()
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Désolée")
                msge.setText("Aucune Location Correspondant Avec Cette Durée !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()  

#****  recherche de location entre deux date ****
          
           
class affiche_loc_deux_date(QDialog):
    def __init__(self):
        super(affiche_loc_deux_date,self).__init__()
        loadUi("affiche_location_entre_deux_date.ui",self)
        self.retour.clicked.connect(self.Return)
        self.afficher.clicked.connect(self.find)    
    def Return(self):
        widget.setCurrentIndex(0)
    def find(self):
        global location
        l=[]
        ok=0
        ok1=0
        date=self.dateEdit.date()
        date1=self.dateEdit_2.date()
        
        heure=self.timeEdit.time()
        heure1=self.timeEdit_2.time()
        #lire date 1
        d1=datetime.date(date.year(),date.month(),date.day())
        #lire date 2
        d2=datetime.date(date1.year(),date1.month(),date1.day())
        #lire heure 1
        hm=datetime.time(heure.hour(),heure.minute())
        #lire heure 2
        hm2=datetime.time(heure1.hour(),heure1.minute())
        if location == {} :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Attention")
            msge.setText("Le Dictionnaire Du Location Est Dejà Vide !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
            widget.setCurrentIndex(0)
        elif d1>d2:
            ok1=0
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Erreur")
            msge.setText("La Premier Date Doit Etre Inferieur A La Deuxiéme !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()
        elif d1==d2 :
            if hm>hm2 :
                ok1=0
                msge = QtWidgets.QMessageBox()
                msge.setWindowTitle("Erreur")
                msge.setText("La Premier Date Doit Etre Inferieur A La Deuxiéme !!")
                msge.setInformativeText("")
                msge.setIcon(QtWidgets.QMessageBox.Critical)
                x = msge.exec_()
            else :
                for i in location :
                    hayt=location[i][2]
                    hout=location[i][3]
                    
                    if d1==hayt :
                        if hm<=hout<=hm2 :
                            l.append(i)
                            ok=1
        
        else :
            for i in location :
                hayt=location[i][2]
                hout=location[i][3]
                    
                if d1<hayt<d2 :
                    l.append(i)
                    ok=1
                elif d1==hayt :
                    if hm<=hout :
                        l.append(i)
                        ok=1
                elif d2==hayt :
                    if hout<=hm2 :
                        l.append(i)
                        ok=1
            
            
        if ok :
            self.tableWidget.setRowCount(len(l))
            row=0
            for i in l:
                date=location[i][2]
                time=location[i][3]
                a=date.year
                m=date.month
                j=date.day
                h=time.hour
                m1=time.minute
                ch=str(j)+"-"+str(m)+"-"+str(a)
                ch1=str(h)+":"+str(m1)
                self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(i))
                self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(location[i][0]))
                self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(location[i][1]))
                self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(ch))
                self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(ch1))
                self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(location[i][4]))            
                self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(location[i][5]))            
                row=row+1
        elif ok==0 and ok1 :
            msge = QtWidgets.QMessageBox()
            msge.setWindowTitle("Désolée")
            msge.setText("Aucune location Trouvée !!")
            msge.setInformativeText("")
            msge.setIcon(QtWidgets.QMessageBox.Critical)
            x = msge.exec_()

 

    

            
app=QApplication(sys.argv)
mainwindow=main()

widget=QtWidgets.QStackedWidget()
widget.addWidget(mainwindow)
widget.setFixedWidth(900)
widget.setFixedHeight(600)
widget.show()
app.exec_()
